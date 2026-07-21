# -*- coding: utf-8 -*-
"""
Created on Sat Dec  8 15:31:39 2018

@author: initial-h
"""


from __future__ import print_function
import random
import numpy as np
import os
import time
import logging
import multiprocessing as mp
import contextlib
import io
from collections import defaultdict, deque
from game_board import Board,Game
from mcts_pure import MCTSPlayer as MCTS_Pure
from greedy_player import GreedyPlayer
from mcts_alphaZero import MCTSPlayer
from policy_value_net_pytorch import PolicyValueNet
from logging_setup import setup_logging, get_logger, setup_worker_logging
from openpyxl import Workbook, load_workbook
import datetime
# QUAN TRONG: KHONG goi setup_logging() ngay o day nua.
#
# Ly do: multiprocessing voi start method 'spawn' se IMPORT LAI train.py
# trong MOI tien trinh con (de tim lai ham _self_play_worker can pickle).
# Neu setup_logging() nam o cap module (chay ngay khi import) thi MOI
# tien trinh con se tu goi lai no -> tao them 1 file log moi + gan them
# 1 console handler moi -> ket qua la 8-10 tien trinh cung in de/moi
# nuoc di ra chung 1 terminal, chong cheo rat kho doc.
#
# Giai phap: chi cau hinh logging co console (setup_logging) trong
# tien trinh CHINH, bang cach dat trong khoi "if __name__ == '__main__'"
# ben duoi file nay. Khi train.py bi import lai boi tien trinh con,
# __name__ luc do la 'train' (khong phai '__main__') nen khoi nay se
# KHONG chay, va tien trinh con se tu cau hinh logging RIENG cho no
# bang setup_worker_logging() (chi ghi file, khong dung console).
logger = get_logger()


def _self_play_worker(args):
    '''
    Ham nay chay o MOT TIEN TRINH (process) rieng, doc lap voi tien trinh chinh.
    Phai dat o cap MODULE (khong nam trong class) de multiprocessing voi
    start method 'spawn' co the pickle va goi lai duoc trong tien trinh con.

    Moi tien trinh con:
      - Tu tao mot Board/Game rieng cua no
      - Tu load mot ban sao PolicyValueNet tu file model tam (tmp_model_path)
      - Choi TRON VEN mot van tu choi doc lap, khong dung chung bo nho voi
        cac tien trinh khac -> nhieu van co the choi song song an toan.
    '''
    (tmp_model_path, board_width, board_height, n_in_row,
     resnet_block, n_playout, c_puct, use_cuda, seed,
     threads_per_worker) = args

    # === QUAN TRONG: gioi han so luong CPU thread cua tien trinh nay ===
    # Mac dinh, PyTorch tu dat so thread = TOAN BO so nhan CPU cua may,
    # cho MOI tien trinh. Khi chay N tien trinh song song ma khong gioi
    # han, N tien trinh se cung tranh gianh toan bo nhan CPU cua nhau
    # (oversubscription) -> ket qua la CHAM HON nhieu so voi chay tuan tu,
    # (dung 1 nuoc di co the mat vai phut thay vi vai giay).
    # Giai phap: moi tien trinh chi dung 1 (hoac vai) thread rieng, vi
    # ban than da co N tien trinh chay song song roi, khong can moi tien
    # trinh lai da luong nua.
    try:
        import torch
        torch.set_num_threads(max(1, threads_per_worker))
    except ImportError:
        pass
    os.environ.setdefault("OMP_NUM_THREADS", str(max(1, threads_per_worker)))
    os.environ.setdefault("MKL_NUM_THREADS", str(max(1, threads_per_worker)))

    # seed rieng cho tung tien trinh, tranh cac van co nuoc di giong het nhau
    random.seed(seed)
    np.random.seed(seed)

    # cau hinh logging RIENG cho tien trinh con nay:
    # - ghi chi tiet (P, Q, u, tung nuoc di...) ra file rieng cua no
    # - KHONG dung console -> terminal chinh khong bi 8-10 tien trinh
    #   in de len nhau nua
    worker_tag = "pid{}".format(os.getpid())
    setup_worker_logging(worker_tag, level=logging.DEBUG)

    t0 = time.time()

    board = Board(width=board_width, height=board_height, n_in_row=n_in_row)
    game = Game(board)

    # PolicyValueNet.__init__ co nhieu print() cung (building network,
    # cuda available, model loaded ...) -> neu khong chan lai, moi tien
    # trinh con se in lai TOAN BO nhung dong nay moi khi choi 1 van,
    # gay ron console. Ta tam thoi "nuot" stdout trong luc khoi tao model.
    with contextlib.redirect_stdout(io.StringIO()):
        policy_value_net = PolicyValueNet(board_width, board_height,
                                           block=resnet_block,
                                           init_model=tmp_model_path,
                                           cuda=use_cuda)

    mcts_player = MCTSPlayer(policy_value_function=policy_value_net.policy_value_fn_random,
                              action_fc=policy_value_net.action_fc_test,
                              evaluation_fc=policy_value_net.evaluation_fc2_test,
                              c_puct=c_puct,
                              n_playout=n_playout,
                              is_selfplay=True)

    # winner, play_data = game.start_self_play(mcts_player, is_shown=False)
    winner, play_data = game.start_self_play_with_greedy(mcts_player, is_shown=False)
    # play_data la generator/zip -> phai ep ve list truoc khi gui qua pipe
    # ve tien trinh cha (pickle khong the truyen thang generator/zip object)
    play_data = list(play_data)

    # 1 dong tom tat gon gang duy nhat in ra console chinh cho moi van
    # (chi tiet day du van con trong logs/selfplay_workers/worker_pid....log)
    elapsed = time.time() - t0
    print("[worker pid={}] van xong: nguoi thang={}, so nuoc={}, thoi gian={:.1f}s".format(
        os.getpid(), winner, len(play_data), elapsed))

    return winner, play_data

class TrainPipeline():
    def __init__(self, init_model=None,transfer_model=None):
        self.resnet_block = 19  # num of block structures in resnet
        # params of the board and the game
        self.board_width = 15
        self.board_height = 15
        self.n_in_row = 5
        self.board = Board(width=self.board_width,
                           height=self.board_height,
                           n_in_row=self.n_in_row)
        self.game = Game(self.board)
        # training params
        self.learn_rate = 1e-3
        self.n_playout = 1000  # num of simulations for each move
        self.c_puct = 5
        # QUAN TRONG VE RAM: buffer_size = so mau (state, mcts_prob, winner)
        # toi da giu trong bo nho. O board 15x15, moi mau ~9KB, nhan augment x8
        # -> 1,000,000 mau co the ngon 8-9GB RAM CHI RIENG buffer nay, cong them
        # policy_update() con copy them 1 lan nua (np.array(self.data_buffer))
        # de shuffle -> co luc gan nhu GAP DOI dung luong do trong chop nhoang.
        # Tren may RAM nho (vd 16GB), can giam han xuong, khong de mac dinh
        # 1,000,000 nhu ban goc (ban goc vien de danh cho may RAM lon hon nhieu).
        self.buffer_size = 100000 # memory size (giam tu 1,000,000 -> phu hop may RAM ~16GB)
        self.batch_size = 2048  # mini-batch size for training
        self.data_buffer = deque(maxlen=self.buffer_size)
        # so tien trinh (process) chay song song de tu choi.
        # KHONG chi dua vao so nhan CPU (mp.cpu_count()) ma con phai dua vao
        # RAM con trong: moi worker tu load 1 ban resnet-19 rieng + cay MCTS
        # rieng, RAM se la nut that co chai truoc ca CPU tren may RAM nho.
        #
        # CAP NHAT cho may i5-13500 (6 nhan P + 8 nhan E = 14 nhan that,
        # 20 luong) / 16GB RAM / RTX 5060 8GB:
        # - 16GB RAM du cho 6 worker + tien trinh chinh ma khong lo swap
        #   (khac han may cu 8GB RAM, noi 10 worker gay tran RAM).
        # - Neu dung selfplay_worker_cuda=True ben duoi, worker chu yeu
        #   cho GPU tra ket qua (it dung CPU lien tuc) nen van co the
        #   tang worker ma khong lo tranh CPU nhieu nhu khi chay full CPU.
        # - Van chua chac chan RAM/VRAM se du cho toi da may worker, nen
        #   BAT DAU voi 6, theo doi RAM (Task Manager) va VRAM (nvidia-smi)
        #   luc chay, tang dan len 8-10 neu con nhieu du dia.
        self.num_selfplay_workers = min(6, mp.cpu_count())
        # cac tien trinh con self-play co dung GPU hay khong.
        # Neu may co GPU ROI (dedicated) dang RANH (nhu GPU 1 trong Task Manager
        # cua ban dang 0%), nen tan dung no: chuyen tinh toan tu CPU/RAM he thong
        # sang VRAM cua GPU, vua nhanh hon (conv net chay tren GPU nhanh hon CPU
        # rat nhieu lan) vua GIAM tai cho RAM he thong.
        # LUU Y: van gioi han so worker dung GPU cho hop VRAM (vd RTX ~8GB thi
        # 2-3 worker la an toan; qua nhieu worker cung tranh VRAM se bi loi/cham).
        #
        # CAP NHAT: may nay co RTX 5060 8GB VRAM (thay vi GTX 1650 4GB truoc
        # do qua chat de dung). Bat GPU cho self-play workers de tan dung -
        # 6 worker cung load resnet-19 (moi ban ~vai tram MB) nen van con
        # nhieu du dia trong 8GB VRAM. Neu gap loi "CUDA out of memory",
        # giam self.num_selfplay_workers xuong hoac dat lai False.
        self.selfplay_worker_cuda = True
        # play n games for each network training
        # nen dat >= num_selfplay_workers de tan dung het cac tien trinh song song
        self.play_batch_size = 200
        self.check_freq = 1
        self.game_batch_num = 1 # total game to train
        self.best_win_ratio = 0.0
        # num of simulations used for the pure mcts, which is used as
        # the opponent to evaluate the trained policy
        self.pure_mcts_playout_num = 100
        if (init_model is not None) and os.path.exists(init_model+'.index'):
            # start training from an initial policy-value net
            self.policy_value_net = PolicyValueNet(self.board_width,self.board_height,block=self.resnet_block,init_model=init_model,cuda=True)
        elif (transfer_model is not None) and os.path.exists(transfer_model+'.index'):
            # start training from a pre-trained policy-value net
            self.policy_value_net = PolicyValueNet(self.board_width,self.board_height,block=self.resnet_block,transfer_model=transfer_model,cuda=True)
        else:
            # start training from a new policy-value net
            self.policy_value_net = PolicyValueNet(self.board_width,self.board_height,block=self.resnet_block,cuda=True)

        self.mcts_player = MCTSPlayer(policy_value_function=self.policy_value_net.policy_value_fn_random,
                                       action_fc=self.policy_value_net.action_fc_test,
                                       evaluation_fc=self.policy_value_net.evaluation_fc2_test,
                                       c_puct=self.c_puct,
                                       n_playout=self.n_playout,
                                       is_selfplay=True)
    def init_excel_logs(self):

        # ---------- Training log ----------
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.train_excel = "excel/training_log_{}.xlsx".format(ts) #"training_log.xlsx"

        if not os.path.exists(self.train_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = "Training"

            ws.append([
                "Batch",
                "Episode Length",
                "Replay Buffer",
                "Loss",
                "Entropy",
                "Collect Time(s)",
                "Train Time(s)",
                "Elapsed Time(h)"
            ])

            wb.save(self.train_excel)

        # ---------- Evaluation log ----------
        self.eval_excel = "excel/evaluation_log_{}.xlsx".format(ts) #"evaluation_log.xlsx"

        if not os.path.exists(self.eval_excel):
            wb = Workbook()
            ws = wb.active
            ws.title = "Evaluation"

            ws.append([
                "Checkpoint",
                "Eval Games",
                "Opponent",
                "AlphaZero Playout",
                "Opponent Playout",
                "Win",
                "Lose",
                "Draw",
                "Win Rate",
                "Best Win Rate",
                "Best Updated",
                "Evaluation Time(s)"
            ])

            wb.save(self.eval_excel)
    def save_training_log(self,
                      batch,
                      loss,
                      entropy,
                      collect_time,
                      train_time,
                      elapsed_time):

        wb = load_workbook(self.train_excel)
        ws = wb["Training"]

        ws.append([
            batch,
            self.episode_len,
            len(self.data_buffer),
            float(loss),
            float(entropy),
            round(collect_time,2),
            round(train_time,2),
            round(elapsed_time/3600,3)
        ])

        wb.save(self.train_excel)
    def save_evaluation_log(self,
                        checkpoint,
                        n_games,
                        opponent,
                        win_cnt,
                        win_ratio,
                        best_updated,
                        eval_time):

        wb = load_workbook(self.eval_excel)
        ws = wb["Evaluation"]

        ws.append([
            checkpoint,
            n_games,
            opponent,
            self.n_playout,
            self.pure_mcts_playout_num if opponent=="PureMCTS" else "-",
            win_cnt[1],
            win_cnt[2],
            win_cnt[-1],
            round(win_ratio,4),
            round(self.best_win_ratio,4),
            "Yes" if best_updated else "No",
            round(eval_time,2)
        ])

        wb.save(self.eval_excel)
    def get_equi_data(self, play_data):
        '''
        augment the data set by rotation and flipping
        play_data: [(state, mcts_prob, winner_z), ..., ...]
        '''
        extend_data = []
        for state, mcts_porb, winner in play_data:
            for i in [1, 2, 3, 4]:
                # rotate counterclockwise
                equi_state = np.array([np.rot90(s, i) for s in state])
                #rotate counterclockwise 90*i
                equi_mcts_prob = np.rot90(np.flipud(
                    mcts_porb.reshape(self.board_height, self.board_width)), i)
                #np.flipud like A[::-1,...]
                #https://docs.scipy.org/doc/numpy-1.6.0/reference/generated/numpy.flipud.html
                # change the reshaped numpy
                # 0,1,2,
                # 3,4,5,
                # 6,7,8,
                # as
                # 6 7 8
                # 3 4 5
                # 0 1 2
                extend_data.append((equi_state,
                                    np.flipud(equi_mcts_prob).flatten(),
                                    winner))
                # flip horizontally
                equi_state = np.array([np.fliplr(s) for s in equi_state])
                #这个np.fliplr like m[:, ::-1]
                #https://docs.scipy.org/doc/numpy/reference/generated/numpy.fliplr.html
                equi_mcts_prob = np.fliplr(equi_mcts_prob)
                extend_data.append((equi_state,
                                    np.flipud(equi_mcts_prob).flatten(),
                                    winner))
        return extend_data

    def collect_selfplay_data(self, n_games=1):
        '''
        collect self-play data for training

        Ban song song: thay vi "for i in range(n_games): choi tung van mot",
        ta phong (spawn) toi da self.num_selfplay_workers tien trinh, moi
        tien trinh choi doc lap 1 van, roi gom ket qua ve. Vi cac van tu
        choi hoan toan doc lap voi nhau (khong van nao phu thuoc ket qua
        van kia), day la bai toan "embarrassingly parallel" -> chi can
        chia viec ra nhieu tien trinh la du, khong can quy hoach dong.
        '''
        # 1) luu model hien tai ra 1 file tam de cac tien trinh con doc lai
        #    (khong the truyen thang object PolicyValueNet/torch qua
        #    multiprocessing, nen phai luu ra file roi cho worker tu load)
        tmp_model_path = os.path.join('tmp', 'selfplay_worker_policy.model')
        if not os.path.exists('tmp'):
            os.makedirs('tmp')
        self.policy_value_net.save_model(tmp_model_path)

        n_workers = max(1, min(n_games, self.num_selfplay_workers))

        # chia deu so nhan CPU cho cac worker, moi worker chi duoc dung
        # phan cua no thoi (xem giai thich chi tiet trong _self_play_worker)
        # -> tranh N tien trinh cung tranh gianh TOAN BO CPU cua nhau
        total_cpu = mp.cpu_count()
        threads_per_worker = max(1, total_cpu // n_workers)

        # 2) chuan bi tham so cho tung van (moi van 1 seed rieng)
        worker_args = [
            (tmp_model_path, self.board_width, self.board_height, self.n_in_row,
             self.resnet_block, self.n_playout, self.c_puct,
             self.selfplay_worker_cuda, random.randint(0, 2 ** 31 - 1),
             threads_per_worker)
            for _ in range(n_games)
        ]

        # 3) chay song song. dung context 'spawn' de an toan voi CUDA/torch
        #    (fork mac dinh tren Linux co the gay loi khi tien trinh cha
        #    da khoi tao CUDA truoc do)
        ctx = mp.get_context('spawn')
        with ctx.Pool(processes=n_workers) as pool:
            results = pool.map(_self_play_worker, worker_args)

        # 4) gom ket qua tu tat ca cac van (nhu vong lap cu, chi khac la
        #    du lieu da duoc choi song song thay vi tuan tu)
        for winner, play_data in results:
            play_data = list(play_data)[:]
            self.episode_len = len(play_data)
            # augment the data
            play_data = self.get_equi_data(play_data)
            self.data_buffer.extend(play_data)

    def policy_update(self):
        '''
        update the policy-value net
        '''
        # play_data: [(state, mcts_prob, winner_z), ..., ...]
        # train an epoch

        tmp_buffer = np.array(self.data_buffer)
        np.random.shuffle(tmp_buffer)
        steps = len(tmp_buffer)//self.batch_size
        print('tmp buffer: {}, steps: {}'.format(len(tmp_buffer),steps))
        for i in range(steps):
            mini_batch = tmp_buffer[i*self.batch_size:(i+1)*self.batch_size]
            state_batch = [data[0] for data in mini_batch]
            mcts_probs_batch = [data[1] for data in mini_batch]
            winner_batch = [data[2] for data in mini_batch]

            old_probs, old_v = self.policy_value_net.policy_value(state_batch=state_batch,
                                                                  actin_fc=self.policy_value_net.action_fc_test,
                                                                  evaluation_fc=self.policy_value_net.evaluation_fc2_test)
            loss, entropy = self.policy_value_net.train_step(state_batch,
                                                             mcts_probs_batch,
                                                             winner_batch,
                                                             self.learn_rate)
            new_probs, new_v = self.policy_value_net.policy_value(state_batch=state_batch,
                                                                  actin_fc=self.policy_value_net.action_fc_test,
                                                                  evaluation_fc=self.policy_value_net.evaluation_fc2_test)
            kl = np.mean(np.sum(old_probs * (
                    np.log(old_probs + 1e-10) - np.log(new_probs + 1e-10)),
                    axis=1)
            )

            explained_var_old = (1 -
                                 np.var(np.array(winner_batch) - old_v.flatten()) /
                                 np.var(np.array(winner_batch)))
            explained_var_new = (1 -
                                 np.var(np.array(winner_batch) - new_v.flatten()) /
                                 np.var(np.array(winner_batch)))

            if steps<10 or (i%(steps//10)==0):
                # print some information, not too much
                logger.info('batch: %d, length: %d, kl:%.5f, loss:%s, entropy:%s, '
                            'explained_var_old:%.3f, explained_var_new:%.3f',
                            i, len(mini_batch), kl, loss, entropy,
                            explained_var_old, explained_var_new)

        return loss, entropy

    def policy_evaluate(self, n_games=10):
        '''
        Evaluate the trained policy by playing against the pure MCTS player
        Note: this is only for monitoring the progress of training
        '''
        current_mcts_player = MCTSPlayer(policy_value_function=self.policy_value_net.policy_value_fn_random,
                                       action_fc=self.policy_value_net.action_fc_test,
                                       evaluation_fc=self.policy_value_net.evaluation_fc2_test,
                                       c_puct=5,
                                       n_playout=1000,
                                       is_selfplay=False)

        # test_player = MCTS_Pure(c_puct=5,
        #                         n_playout=self.pure_mcts_playout_num)
        test_player = GreedyPlayer()
        # test_player = MCTSPlayer(policy_value_function=self.policy_value_net.policy_value_fn_random,
        #                                action_fc=self.policy_value_net.action_fc_test,
        #                                evaluation_fc=self.policy_value_net.evaluation_fc2_test,
        #                                c_puct=5,
        #                                n_playout=400,
        #                                is_selfplay=False)
        
        win_cnt = defaultdict(int)
        for i in range(n_games):
            print('playing game %d.' % i)
            winner = self.game.start_play(player1=current_mcts_player,
                                          player2=test_player,
                                          start_player=i % 2,
                                          is_shown=1,
                                          print_prob=False)
            win_cnt[winner] += 1
        win_ratio = 1.0*(win_cnt[1] + 0.5*win_cnt[-1]) / n_games
        logger.info("[DANH GIA]:, thang:%d, thua:%d, hoa:%d, win_ratio:%.3f",
                     win_cnt[1], win_cnt[2], win_cnt[-1], win_ratio)
        return win_ratio, win_cnt

    def run(self):
        '''
        run the training pipeline
        '''
        # make dirs first
        if not os.path.exists('tmp'):
            os.makedirs('tmp')
        if not os.path.exists('model'):
            os.makedirs('model')
        self.init_excel_logs()

        # record time for each part
        start_time = time.time()
        collect_data_time = 0
        train_data_time = 0
        evaluate_time = 0

        try:
            for i in range(self.game_batch_num):
                # collect self-play data
                collect_data_start_time = time.time()
                self.collect_selfplay_data(self.play_batch_size)
                collect_data_time += time.time()-collect_data_start_time
                logger.info("===== batch tu choi thu %d, so nuoc di trong van (episode_len)=%d =====",
                            i+1, self.episode_len)

                if len(self.data_buffer) > self.batch_size*5:
                    # train collected data
                    train_data_start_time = time.time()
                    loss, entropy = self.policy_update()
                    train_data_time += time.time()-train_data_start_time
                    self.save_training_log(
                        batch=i+1,
                        loss=loss,
                        entropy=entropy,
                        collect_time=collect_data_time,
                        train_time=train_data_time,
                        elapsed_time=time.time()-start_time
                    )
                    # print some training information
                    logger.info('thoi gian da chay: %.3f gio', (time.time() - start_time) / 3600)
                    logger.info('collect_data_time: %.3f h, train_data_time: %.3f h, evaluate_time: %.3f h',
                                collect_data_time / 3600, train_data_time / 3600, evaluate_time / 3600)

                if (i+1) % self.check_freq == 0 :

                    # save current model for evaluating
                    self.policy_value_net.save_model('tmp/current_policy.model')
                    if (i+1) % self.check_freq == 0:
                    # if (i+1) % (self.check_freq*2) == 0:
                        print("current self-play batch: {}".format(i + 1))
                        evaluate_start_time = time.time()

                        # evaluate current model
                        best_updated = False
                        print('Đang đánh giá model với 100 van đánh:')
                        win_ratio, win_cnt = self.policy_evaluate(n_games=100)
                        evaluate_time += time.time()-evaluate_start_time
                        if win_ratio > self.best_win_ratio:
                            # save best model
                            # print("New best policy!!!!!!!!")
                            best_updated = True
                            self.best_win_ratio = win_ratio
                            self.policy_value_net.save_model('model/best_policy.model')
                            
                            # if (self.best_win_ratio == 1.0 and self.pure_mcts_playout_num < 5000):
                            #     # increase playout num and  reset the win ratio
                            #     self.pure_mcts_playout_num += 100
                            #     self.best_win_ratio = 0.0
                            # if self.pure_mcts_playout_num ==5000:
                            #     # reset mcts pure playout num
                            #     self.pure_mcts_playout_num = 1000
                            #     self.best_win_ratio = 0.0
                        self.save_evaluation_log(
                                checkpoint=i+1,
                                n_games=100,
                                opponent="AlphaZero O",
                                win_cnt=win_cnt,
                                win_ratio=win_ratio,
                                best_updated=best_updated,
                                eval_time=evaluate_time
                            )
        except KeyboardInterrupt:
            print('\n\rquit')

if __name__ == '__main__':
    # Cau hinh logging co CONSOLE chi o day - tien trinh CHINH.
    # Khi cac tien trinh con (worker) import lai file nay de goi
    # _self_play_worker, __name__ cua chung la 'train' chu khong phai
    # '__main__', nen khoi nay se KHONG chay lai trong worker.
    # - level=logging.DEBUG        : file log chinh chua chi tiet playout cua MCTS o tien trinh chinh (vd policy_evaluate)
    # - console_level=logging.INFO : man hinh chi hien tom tat batch/epoch, khong bi tran boi tung nuoc di cua tung van song song
    setup_logging(level=logging.INFO, console_level=logging.INFO)
    logger = get_logger()

    # training_pipeline = TrainPipeline(init_model='model/best_policy.model',transfer_model=None)
    # training_pipeline = TrainPipeline(init_model=None, transfer_model='transfer_model/best_policy.model')
    training_pipeline = TrainPipeline()
    training_pipeline.run()